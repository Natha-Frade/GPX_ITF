# -*- coding: utf-8 -*-
"""
Validação de Campo — GPX IMTRAFF
Motor de validação por MARCOS QUILOMÉTRICOS.

Como funciona
-------------
1. O KMZ traz os marcos de km da(s) rodovia(s). Os marcos são encadeados
   numa ROTA contínua por rodovia — a numeração reinicia na divisa entre
   estados, então a rota costura as séries pelas pontas.
2. A planilha de filmagem define os cortes esperados. Só as colunas
   B (trecho/rodovia), C (pista/faixa), D (km inicial) e E (km final)
   são lidas.
3. Cada GPX é projetado na rota: vira um intervalo de quilometragem.
4. Cruzando os dois, sai quanto de cada corte foi coberto, por câmera, e
   onde ficaram os buracos.

O eixo linear da rota é a distância acumulada entre os marcos, não a
diferença de numeração — a numeração tem saltos e trechos que não medem
1 km de verdade. Sem dependências além de openpyxl; o resto é stdlib.
"""
import io
import re
import math
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

R_TERRA = 6371000.0
KML_NS = '{http://www.opengis.net/kml/2.2}'
GPX_NS = '{http://www.topografix.com/GPX/1/1}'

# ---------------------------------------------------------------- parâmetros
DIST_MAX_ROTA = 120        # m — ponto além disso não conta como "na rodovia"
V_MAX = 60.0               # m/s (216 km/h) — acima disso é salto de GPS
SALTO_MIN = 1500.0         # m — costura a oscilação entre pistas na projeção
SALTO_MAX = 20000.0        # m — teto do que um intervalo de tempo pode cobrir
TOL_FECHADO = 0.99         # ≥99% do corte coberto = FECHADO
TOL_PARCIAL = 0.05         # <5% = praticamente sem cobertura
BURACO_MIN = 200           # m — buraco menor que isso é ruído, não pendência
JUNCAO_MAX = 3000          # m — marco da pista oposta só vale se estiver perto
FAIXAS_NA_OBS = 6          # quantas faixas de km cabem na observação


# ---------------------------------------------------------------- geometria
def _dist_m(a, b):
    """Distância equiretangular (suficiente p/ escalas < ~50 km)."""
    x = math.radians(b[1] - a[1]) * math.cos(math.radians((a[0] + b[0]) / 2))
    y = math.radians(b[0] - a[0])
    return math.hypot(x, y) * R_TERRA


def fmt_km(km):
    """7.5 -> '007+500'"""
    if km is None:
        return None
    inteiro = int(math.floor(km))
    metros = int(round((km - inteiro) * 1000))
    if metros >= 1000:
        inteiro, metros = inteiro + 1, 0
    return f"{inteiro:03d}+{metros:03d}"


def parse_km_planilha(txt):
    """'215+000' -> 215.0 ; '007+500' -> 7.5 ; '215' -> 215.0"""
    if txt is None:
        return None
    s = str(txt).strip()
    m = re.match(r'^(\d{1,4})\s*\+\s*(\d{1,3})$', s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 1000.0
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        return None


def _normalizar_rodovia(txt):
    """'BR 116' / 'br-116' -> 'BR-116'. None se não parecer rodovia."""
    if not txt:
        return None
    s = re.sub(r'\s+', '', str(txt)).upper()
    m = re.match(r'^(BR)-?(\d{3})$', s)
    return f"{m.group(1)}-{m.group(2)}" if m else None


# ------------------------------------------------------- KMZ: marcos de km
def parse_kmz(dados: bytes):
    """KMZ/KML dos marcos -> {rodovia: [{km, lat, lon, alt}]}.

    `alt` marca o marco da pista oposta (nome tipo "224 Norte"). A rodovia
    vem do nome da pasta ("Marcos_km_BR116").
    """
    raw = dados
    if raw[:2] == b'PK':
        z = zipfile.ZipFile(io.BytesIO(raw))
        nomes = [n for n in z.namelist() if n.lower().endswith('.kml')]
        if not nomes:
            raise ValueError("KMZ sem .kml dentro.")
        raw = z.read(nomes[0])
    root = ET.fromstring(raw.decode('utf-8', 'replace'))

    rodovias = {}
    for folder in root.iter(KML_NS + 'Folder'):
        fn = folder.find(KML_NS + 'name')
        texto = fn.text if fn is not None and fn.text else ''
        achou = re.search(r'BR[-\s]?\d{3}', texto, re.I)
        rod = _normalizar_rodovia(achou.group(0)) if achou else None
        if not rod:
            continue
        marcos = []
        for pm in folder.findall('.//' + KML_NS + 'Placemark'):
            n = pm.find(KML_NS + 'name')
            c = pm.find('.//' + KML_NS + 'Point/' + KML_NS + 'coordinates')
            if n is None or c is None or not n.text or not c.text:
                continue
            mm = re.match(r'^\s*(\d{1,4})\s*(.*)$', n.text.strip())
            if not mm:
                continue
            try:
                lon, lat = [float(v) for v in c.text.strip().split(',')[:2]]
            except ValueError:
                continue
            marcos.append({'km': int(mm.group(1)), 'lat': lat, 'lon': lon,
                           'alt': bool(mm.group(2).strip())})
        if marcos:
            rodovias.setdefault(rod, []).extend(marcos)
    return rodovias


def montar_rota(marcos):
    """Encadeia os marcos numa rota contínua.

    A numeração reinicia na divisa, então primeiro montamos cadeias por
    continuidade geográfica (km a km) e depois unimos as cadeias pelas
    pontas mais próximas. Devolve [{km, lat, lon, off}] com `off` = metros
    acumulados desde o início da rota.
    """
    porkm = {}
    for m in marcos:
        if not m['alt']:
            porkm.setdefault(m['km'], []).append(m)

    cadeias = []
    for km in sorted(porkm):
        for m in porkm[km]:
            melhor, melhor_d = None, float('inf')
            for c in cadeias:
                ult = c[-1]
                if km - ult['km'] > 3:          # buraco grande na numeração
                    continue
                d = _dist_m((ult['lat'], ult['lon']), (m['lat'], m['lon']))
                if d < melhor_d:
                    melhor, melhor_d = c, d
            # marcos consecutivos ficam a ~1 km; a folga cresce com o pulo
            folga = 4000 * max(1, km - melhor[-1]['km']) if melhor else 0
            if melhor is not None and melhor_d < folga:
                melhor.append(m)
            else:
                cadeias.append([m])

    cadeias = [c for c in cadeias if len(c) >= 2]
    if not cadeias:
        return []

    rota = cadeias.pop(0)
    while cadeias:
        alvo, melhor_d, virar_rota, virar_alvo = None, float('inf'), False, False
        for c in cadeias:
            for vr in (False, True):
                pa = rota[0] if vr else rota[-1]
                for vc in (False, True):
                    pb = c[-1] if vc else c[0]
                    d = _dist_m((pa['lat'], pa['lon']), (pb['lat'], pb['lon']))
                    if d < melhor_d:
                        alvo, melhor_d, virar_rota, virar_alvo = c, d, vr, vc
        cadeias.remove(alvo)
        if melhor_d > 5000:        # cadeia solta: não faz parte desta rota
            continue
        if virar_rota:
            rota.reverse()
        rota += list(reversed(alvo)) if virar_alvo else alvo

    saida, off = [dict(rota[0], off=0.0)], 0.0
    for i in range(1, len(rota)):
        off += _dist_m((rota[i - 1]['lat'], rota[i - 1]['lon']),
                       (rota[i]['lat'], rota[i]['lon']))
        saida.append(dict(rota[i], off=off))
    return saida


def montar_variantes(rota, marcos):
    """Pista oposta (marcos "224 Norte"). A numeração é a mesma, então a
    variante compartilha o eixo de offsets — só troca a geometria onde as
    pistas se separam."""
    alt = {}
    for m in marcos:
        if m['alt']:
            alt.setdefault(m['km'], []).append(m)
    if not alt:
        return []
    variante, tocou = [], False
    for m in rota:
        # o mesmo número de km aparece em séries diferentes da rodovia, então
        # o marco da pista oposta só vale se estiver ao lado deste — senão
        # clonaríamos a geometria de uma série em cima da outra.
        cand = [a for a in alt.get(m['km'], [])
                if _dist_m((m['lat'], m['lon']), (a['lat'], a['lon'])) < JUNCAO_MAX]
        if cand:
            a = min(cand, key=lambda x: _dist_m((m['lat'], m['lon']),
                                                (x['lat'], x['lon'])))
            variante.append({'km': m['km'], 'lat': a['lat'], 'lon': a['lon'],
                             'off': m['off']})
            tocou = True
        else:
            variante.append(m)
    return [variante] if tocou else []


def montar_rotas(rodovias):
    """{rodovia: marcos} -> {rodovia: [rota_principal, variante...]}"""
    rotas = {}
    for rod, marcos in rodovias.items():
        principal = montar_rota(marcos)
        if principal:
            rotas[rod] = [principal] + montar_variantes(principal, marcos)
    return rotas


# ------------------------------------------------------------- projeção
def projetar(rota, lat, lon):
    """Ponto -> (offset_m na rota, distância_m até a rota)."""
    melhor_i, melhor_d = None, float('inf')
    for i, m in enumerate(rota):
        d = _dist_m((m['lat'], m['lon']), (lat, lon))
        if d < melhor_d:
            melhor_i, melhor_d = i, d
    if melhor_i is None:
        return None

    melhor = None
    for j in (melhor_i - 1, melhor_i):
        if j < 0 or j + 1 >= len(rota):
            continue
        a, b = rota[j], rota[j + 1]
        seg = b['off'] - a['off']
        if seg <= 0:
            continue
        cos = math.cos(math.radians(a['lat']))
        bx = math.radians(b['lon'] - a['lon']) * cos * R_TERRA
        by = math.radians(b['lat'] - a['lat']) * R_TERRA
        px = math.radians(lon - a['lon']) * cos * R_TERRA
        py = math.radians(lat - a['lat']) * R_TERRA
        L2 = bx * bx + by * by
        t = 0.0 if L2 == 0 else max(0.0, min(1.0, (px * bx + py * by) / L2))
        d = math.hypot(px - bx * t, py - by * t)
        if melhor is None or d < melhor[1]:
            melhor = (a['off'] + seg * t, d)
    return melhor if melhor is not None else (rota[melhor_i]['off'], melhor_d)


def km_do_offset(rota, off):
    """Offset (m) -> km na numeração dos marcos."""
    if not rota:
        return None
    if off <= rota[0]['off']:
        return float(rota[0]['km'])
    for i in range(1, len(rota)):
        if off <= rota[i]['off']:
            a, b = rota[i - 1], rota[i]
            seg = b['off'] - a['off']
            t = 0.0 if seg <= 0 else (off - a['off']) / seg
            dkm = b['km'] - a['km']
            if abs(dkm) > 3:            # divisa: a numeração reinicia
                return float(b['km']) if t > 0.5 else float(a['km'])
            return a['km'] + dkm * t
    return float(rota[-1]['km'])


def offsets_do_km(rota, km_alvo):
    """Todos os offsets onde a numeração passa por `km_alvo`."""
    saida = []
    for i in range(1, len(rota)):
        a, b = rota[i - 1], rota[i]
        dkm = b['km'] - a['km']
        if abs(dkm) > 3:
            continue
        if min(a['km'], b['km']) <= km_alvo <= max(a['km'], b['km']):
            seg = b['off'] - a['off']
            t = 0.0 if dkm == 0 else (km_alvo - a['km']) / dkm
            saida.append(a['off'] + seg * t)
    saida.sort()
    limpo = []
    for o in saida:
        if not limpo or o - limpo[-1] > 500:
            limpo.append(o)
    return limpo


# ------------------------------------------------- planilha de filmagem
def parse_planilha(dados: bytes):
    """Lê a planilha de filmagem. Só colunas B, C, D e E.

    B = trecho (rodovia) | C = pista/faixa | D = km inicial | E = km final
    A coluna A, quando existe, vira o rótulo da linha.
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(dados), data_only=True, read_only=True)
    cortes = []
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not row or len(row) < 5:
                continue
            rod = _normalizar_rodovia(row[1])
            if not rod:
                continue
            km_ini = parse_km_planilha(row[3])
            km_fim = parse_km_planilha(row[4])
            if km_ini is None or km_fim is None:
                continue
            cortes.append({
                'aba': ws.title,
                'linha': i,
                'nome': str(row[0]).strip() if row[0] else f'Linha {i}',
                'rodovia': rod,
                'pista': str(row[2]).strip() if row[2] else '',
                'km_ini': km_ini,
                'km_fim': km_fim,
                'km_ini_str': fmt_km(km_ini),
                'km_fim_str': fmt_km(km_fim),
            })
    wb.close()
    return cortes


def localizar_corte(rota, corte):
    """(offset_ini, offset_fim) do corte na rota.

    O mesmo km aparece mais de uma vez quando a numeração reinicia; o par
    certo é o menor trecho contíguo que liga os dois extremos.
    """
    ini = offsets_do_km(rota, corte['km_ini'])
    fim = offsets_do_km(rota, corte['km_fim'])
    if not ini or not fim:
        return None
    return min(((a, b) for a in ini for b in fim),
               key=lambda p: abs(p[1] - p[0]))


# ------------------------------------------------------------------- GPX
def parse_gpx(dados: bytes):
    root = ET.fromstring(dados.decode('utf-8', 'replace'))
    ns = GPX_NS if root.tag.startswith(GPX_NS) else ''
    pontos = []
    for tp in root.iter(ns + 'trkpt'):
        t = tp.find(ns + 'time')
        quando = None
        if t is not None and t.text:
            try:
                quando = datetime.strptime(t.text.strip()[:19],
                                           '%Y-%m-%dT%H:%M:%S')
            except ValueError:
                pass
        try:
            pontos.append((float(tp.get('lat')), float(tp.get('lon')), quando))
        except (TypeError, ValueError):
            continue
    return {'points': pontos}


def extrair_gpx_de_upload(nome_arq, dados, camera_pai=None):
    """Abre .gpx ou .zip (recursivo). A câmera vem do nome da pasta/zip."""
    saida = []
    low = (nome_arq or '').lower()

    def camera_de(texto, atual):
        if re.search(r'intern', texto, re.I):
            return 'INTERNO'
        if re.search(r'extern', texto, re.I):
            return 'EXTERNO'
        return atual

    if low.endswith('.zip'):
        cam = camera_de(nome_arq.rsplit('/', 1)[-1], camera_pai)
        try:
            z = zipfile.ZipFile(io.BytesIO(dados))
        except zipfile.BadZipFile:
            return saida
        for n in z.namelist():
            if not n.endswith('/'):
                saida += extrair_gpx_de_upload(n, z.read(n), cam)
        return saida

    if low.endswith('.gpx'):
        saida.append({'arquivo': nome_arq.rsplit('/', 1)[-1],
                      'camera': camera_de(nome_arq, camera_pai) or 'N/D',
                      'dados': dados})
    return saida


def nome_video(arquivo):
    """'GX011550-002_1_GPS5.gpx' -> 'GX011550'"""
    m = re.search(r'(G[XHPS]\d{6})', arquivo, re.I)
    return m.group(1).upper() if m else re.sub(r'\.gpx$', '', arquivo,
                                               flags=re.I)


# -------------------------------------------------------------- cobertura
def _uniao(ivs):
    if not ivs:
        return []
    ivs = sorted(ivs)
    saida = [list(ivs[0])]
    for a, b in ivs[1:]:
        if a <= saida[-1][1] + 1:
            saida[-1][1] = max(saida[-1][1], b)
        else:
            saida.append([a, b])
    return saida


def _intervalos(offs, tempos):
    """Offsets na ordem de gravação -> intervalos percorridos.

    Pontos consecutivos no tempo foram percorridos de fato, então cobrimos
    par a par. O limite do que se aceita é físico: o que cabe no intervalo
    de tempo a V_MAX. Isso costura a oscilação da projeção onde as pistas
    se separam, sem inventar cobertura em cima de um salto de GPS.
    """
    segs = []
    for i, (a, b) in enumerate(zip(offs, offs[1:])):
        limite = SALTO_MIN
        ta, tb = tempos[i], tempos[i + 1]
        if ta and tb:
            dt = abs((tb - ta).total_seconds())
            limite = max(SALTO_MIN, min(V_MAX * dt, SALTO_MAX))
        if abs(b - a) <= limite:
            segs.append((min(a, b), max(a, b)))
    return [tuple(iv) for iv in _uniao(segs)]


def _medir(ivs, a, b):
    """Quanto de [a,b] os intervalos cobrem."""
    return sum(max(0.0, min(b, y) - max(a, x)) for x, y in _uniao(ivs))


def _buracos(ivs, a, b):
    """Pedaços de [a,b] que ficaram sem cobertura."""
    faltando, cursor = [], a
    for x, y in _uniao(ivs):
        if y <= a or x >= b:
            continue
        if x > cursor:
            faltando.append((cursor, min(x, b)))
        cursor = max(cursor, min(y, b))
    if cursor < b:
        faltando.append((cursor, b))
    return [f for f in faltando if f[1] - f[0] >= BURACO_MIN]


def _cobertos(ivs, a, b):
    """Pedaços de [a,b] efetivamente cobertos."""
    return [(max(a, x), min(b, y)) for x, y in _uniao(ivs)
            if min(b, y) - max(a, x) >= BURACO_MIN]


def _faixas_km(rota, faixas, off_ini, off_fim, limite=FAIXAS_NA_OBS):
    """Faixas de offset -> texto 'km 215+000 → 240+023; km ...'.

    As faixas saem na ordem de leitura do corte: começam pelo km inicial e
    seguem no sentido dele, mesmo quando a numeração é decrescente.
    """
    if not faixas:
        return ''
    invertido = off_fim < off_ini
    ordenadas = sorted(faixas, key=lambda f: abs(f[0] - off_ini),
                       reverse=False)
    textos = []
    for x, y in ordenadas[:limite]:
        de, ate = (y, x) if invertido else (x, y)
        textos.append(f'{fmt_km(km_do_offset(rota, de))} → '
                      f'{fmt_km(km_do_offset(rota, ate))}')
    if len(ordenadas) > limite:
        textos.append(f'+{len(ordenadas) - limite} trecho(s)')
    return '; '.join(textos)


def projetar_trilha(rotas, pontos, passo=1):
    """Projeta a trilha; devolve a rodovia que melhor explica o percurso."""
    melhor = None
    for rod, variantes in rotas.items():
        offs, desvios, tempos = [], [], []
        for (lat, lon, t) in pontos[::passo]:
            p = None
            for rota in variantes:
                q = projetar(rota, lat, lon)
                if q is not None and (p is None or q[1] < p[1]):
                    p = q
            if p is None:
                continue
            offs.append(p[0])
            desvios.append(p[1])
            tempos.append(t)
        if len(offs) < 2:
            continue
        score = sum(1 for d in desvios if d < DIST_MAX_ROTA) / len(desvios)
        if melhor is None or score > melhor['score']:
            melhor = {'rodovia': rod, 'score': score, 'offs': offs,
                      'desvios': desvios, 'tempos': tempos,
                      'intervalos': _intervalos(offs, tempos)}
    return melhor


# ------------------------------------------------------------- consultas
def cortes_por_km(cortes, km_ini, km_fim=None, rodovia=None):
    """Cortes da planilha que cobrem um km (ou intervalo)."""
    a = km_ini if km_fim is None else min(km_ini, km_fim)
    b = km_ini if km_fim is None else max(km_ini, km_fim)
    achados = []
    for c in cortes:
        if rodovia and c['rodovia'] != _normalizar_rodovia(rodovia):
            continue
        lo, hi = min(c['km_ini'], c['km_fim']), max(c['km_ini'], c['km_fim'])
        if hi >= a and lo <= b:
            achados.append(c)
    return sorted(achados, key=lambda c: (c['rodovia'],
                                          min(c['km_ini'], c['km_fim'])))


def estimar_km_gpx(pontos, rodovias):
    """Trilha -> (rodovia, km inicial, km final) pelos marcos do KMZ."""
    rotas = montar_rotas(rodovias)
    proj = projetar_trilha(rotas, pontos) if rotas else None
    if proj is None or proj['score'] < 0.2:
        return (None, None, None)
    rota = rotas[proj['rodovia']][0]
    return (proj['rodovia'],
            km_do_offset(rota, proj['offs'][0]),
            km_do_offset(rota, proj['offs'][-1]))


# ---------------------------------------------------------------- validação
def _hora(t):
    return t.strftime('%H:%M:%S') if t else None


def _dia(t):
    return t.strftime('%d-%m') if t else None


def validar_lote(entradas, rodovias, cortes):
    """Cruza os GPX enviados com os cortes esperados.

    entradas: [{arquivo, camera, dados}] — saída de extrair_gpx_de_upload
    rodovias: saída de parse_kmz
    cortes:   saída de parse_planilha
    """
    rotas = montar_rotas(rodovias)
    if not rotas:
        raise ValueError("O KMZ não tem marcos de km suficientes para montar "
                         "a rota da rodovia.")

    # 1) cada GPX vira um intervalo de quilometragem na sua rodovia
    videos, ignorados = [], []
    for ent in entradas:
        try:
            g = parse_gpx(ent['dados'])
        except ET.ParseError as e:
            ignorados.append({'arquivo': ent['arquivo'],
                              'motivo': f'GPX inválido: {e}'})
            continue
        if len(g['points']) < 2:
            ignorados.append({'arquivo': ent['arquivo'],
                              'motivo': 'GPX sem pontos suficientes'})
            continue
        proj = projetar_trilha(rotas, g['points'])
        if proj is None or proj['score'] < 0.2:
            ignorados.append({'arquivo': ent['arquivo'],
                              'motivo': 'trilha não bate com nenhuma rodovia '
                                        'do KMZ'})
            continue
        rota = rotas[proj['rodovia']][0]
        tempos = [t for t in proj['tempos'] if t]
        desvios = sorted(proj['desvios'])
        videos.append({
            'arquivo': ent['arquivo'],
            'video': nome_video(ent['arquivo']),
            'camera': ent.get('camera') or 'N/D',
            'rodovia': proj['rodovia'],
            'intervalos': proj['intervalos'],
            'pontos': len(g['points']),
            'desvio_mediano': round(desvios[len(desvios) // 2], 1),
            'km_ini': fmt_km(km_do_offset(rota, proj['offs'][0])),
            'km_fim': fmt_km(km_do_offset(rota, proj['offs'][-1])),
            'inicio': min(tempos) if tempos else None,
            'fim': max(tempos) if tempos else None,
        })

    # 2) cada corte esperado recebe os vídeos que passam por ele
    linhas, detalhes, pendencias = [], [], []
    for corte in cortes:
        variantes = rotas.get(corte['rodovia'])
        base = dict(corte, cobertura=0.0, cobertura_cam={}, videos={},
                    extensao_km=None)
        if not variantes:
            linhas.append(dict(base, status='SEM REFERÊNCIA',
                               observacao='Rodovia ausente no KMZ'))
            continue
        rota = variantes[0]
        par = localizar_corte(rota, corte)
        if par is None:
            linhas.append(dict(base, status='SEM REFERÊNCIA',
                               observacao='Km fora dos marcos do KMZ'))
            continue
        a, b = min(par), max(par)
        extensao = b - a
        if extensao <= 0:
            linhas.append(dict(base, status='SEM REFERÊNCIA',
                               observacao='Km inicial igual ao final'))
            continue

        porcam = {}
        for v in videos:
            if v['rodovia'] != corte['rodovia']:
                continue
            coberto = _medir(v['intervalos'], a, b)
            if coberto < BURACO_MIN:
                continue
            porcam.setdefault(v['camera'], []).append(v)
            detalhes.append({
                'corte': corte['nome'], 'rodovia': corte['rodovia'],
                'trecho': f"{corte['km_ini_str']} → {corte['km_fim_str']}",
                'camera': v['camera'], 'video': v['video'],
                'arquivo': v['arquivo'],
                'km_ini_video': v['km_ini'], 'km_fim_video': v['km_fim'],
                'km_cobertos': round(coberto / 1000, 3),
                'pct': round(100 * coberto / extensao, 1),
                'desvio_mediano': v['desvio_mediano'],
                'dia': _dia(v['inicio']),
                'hora_ini': _hora(v['inicio']), 'hora_fim': _hora(v['fim']),
            })

        cobertura_cam = {}
        for cam, vs in porcam.items():
            ivs = [iv for v in vs for iv in v['intervalos']]
            cobertura_cam[cam] = _medir(ivs, a, b) / extensao

        todos = [iv for vs in porcam.values() for v in vs
                 for iv in v['intervalos']]
        cobertura = _medir(todos, a, b) / extensao if todos else 0.0
        faltando_cam = [c for c in ('INTERNO', 'EXTERNO')
                        if cobertura_cam.get(c, 0.0) < TOL_FECHADO]

        cobertos = _cobertos(todos, a, b)
        faltantes = _buracos(todos, a, b)
        txt_cob = _faixas_km(rota, cobertos, *par)
        txt_falta = _faixas_km(rota, faltantes, *par)

        if cobertura < TOL_PARCIAL:
            status = 'NÃO FILMADO'
            obs = (f'Nenhum vídeo cobre este corte — faltam os '
                   f'{extensao / 1000:.3f} km: '
                   f'{fmt_km(corte["km_ini"])} → {fmt_km(corte["km_fim"])}')
        elif cobertura >= TOL_FECHADO and not faltando_cam:
            status = 'FECHADO'
            obs = f'100% coberto: {txt_cob}' if txt_cob else None
        else:
            status = 'PARCIAL'
            partes = [f'{cobertura * 100:.1f}% coberto'
                      + (f': {txt_cob}' if txt_cob else '')]
            if faltantes:
                falta_km = sum(y - x for x, y in faltantes) / 1000
                partes.append(f'faltam {falta_km:.3f} km: {txt_falta}')
            for c in faltando_cam:
                if c in porcam:
                    pendente = _faixas_km(
                        rota, _buracos([iv for v in porcam[c]
                                        for iv in v['intervalos']], a, b), *par)
                    partes.append(f'{c.lower()} cobre '
                                  f'{cobertura_cam[c] * 100:.1f}%'
                                  + (f', falta {pendente}' if pendente else ''))
                else:
                    partes.append(f'sem vídeo {c.lower()}')
            obs = '. '.join(partes)

        if status != 'NÃO FILMADO' and any(len(vs) > 1 for vs in porcam.values()):
            extra = 'mais de 1 vídeo por câmera (junção necessária)'
            obs = f'{obs}. {extra}' if obs else extra

        for x, y in _buracos(todos, a, b):
            pendencias.append({
                'corte': corte['nome'], 'rodovia': corte['rodovia'],
                'km_ini': fmt_km(km_do_offset(rota, x)),
                'km_fim': fmt_km(km_do_offset(rota, y)),
                'extensao_km': round((y - x) / 1000, 3),
                'motivo': 'Sem cobertura de nenhuma câmera' if not todos
                          else 'Buraco entre os vídeos enviados',
            })

        ordem = lambda v: v['inicio'] or datetime.min
        linhas.append(dict(
            base, status=status,
            cobertura=round(cobertura, 4),
            extensao_km=round(extensao / 1000, 3),
            cobertura_cam={c: round(v, 4) for c, v in cobertura_cam.items()},
            videos={c: [v['video'] for v in sorted(vs, key=ordem)]
                    for c, vs in porcam.items()},
            observacao=obs))

    usados = {d['arquivo'] for d in detalhes}
    sobrando = [{'arquivo': v['arquivo'], 'video': v['video'],
                 'camera': v['camera'], 'rodovia': v['rodovia'],
                 'km_ini': v['km_ini'], 'km_fim': v['km_fim'],
                 'motivo': 'Trilha fora de todos os cortes da planilha'}
                for v in videos if v['arquivo'] not in usados]

    resumo = {
        'cortes': len(linhas),
        'fechados': sum(1 for l in linhas if l['status'] == 'FECHADO'),
        'parciais': sum(1 for l in linhas if l['status'] == 'PARCIAL'),
        'nao_filmados': sum(1 for l in linhas if l['status'] == 'NÃO FILMADO'),
        'sem_referencia': sum(1 for l in linhas
                              if l['status'] == 'SEM REFERÊNCIA'),
        'gpx_lidos': len(videos),
        'gpx_ignorados': len(ignorados),
        'gpx_sobrando': len(sobrando),
    }
    return {'resumo': resumo, 'linhas': linhas, 'detalhes': detalhes,
            'pendencias': pendencias, 'sobrando': sobrando,
            'ignorados': ignorados,
            'rodovias': {r: {'marcos': len(rotas[r][0]),
                             'extensao_km': round(rotas[r][0][-1]['off'] / 1000, 3)}
                         for r in rotas}}


# ------------------------------------------------------------ exportação
_STATUS_COR = {
    'FECHADO': 'FFC6EFCE',
    'PARCIAL': 'FFFFEB9C',
    'NÃO FILMADO': 'FFFFC7CE',
    'SEM REFERÊNCIA': 'FFD9D9D9',
}


def exportar_xlsx(relatorio, ref_kmz='', ref_plan=''):
    """Relatório -> planilha (RESUMO / DETALHADO / NÃO COBERTOS / CRITÉRIOS)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fundo = PatternFill('solid', fgColor='FF1F3864')
    titulo = Font(bold=True, color='FFFFFFFF')

    def montar(ws, colunas, linhas, larguras):
        ws.append(colunas)
        for c in range(1, len(colunas) + 1):
            cel = ws.cell(row=1, column=c)
            cel.font = titulo
            cel.fill = fundo
            cel.alignment = Alignment(vertical='center', wrap_text=True)
        for linha in linhas:
            ws.append(linha)
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        if linhas:
            ws.auto_filter.ref = ws.dimensions

    # RESUMO
    ws = wb.active
    ws.title = 'RESUMO'
    montar(ws, ['CORTE', 'TRECHO', 'PISTA/FAIXA', 'KM INICIAL', 'KM FINAL',
                'EXTENSÃO (km)', 'VÍDEO(S) INTERNO', 'VÍDEO(S) EXTERNO',
                'COBERTURA (%)', 'STATUS', 'OBSERVAÇÃO'],
           [[l['nome'], l['rodovia'], l.get('pista', ''),
             l['km_ini_str'], l['km_fim_str'], l.get('extensao_km'),
             '\n'.join(l.get('videos', {}).get('INTERNO', [])) or None,
             '\n'.join(l.get('videos', {}).get('EXTERNO', [])) or None,
             round(l['cobertura'] * 100, 1), l['status'], l.get('observacao')]
            for l in relatorio['linhas']],
           [12, 10, 12, 12, 12, 14, 22, 22, 14, 16, 46])
    for i, l in enumerate(relatorio['linhas'], 2):
        cor = _STATUS_COR.get(l['status'])
        if cor:
            ws.cell(row=i, column=10).fill = PatternFill('solid', fgColor=cor)
        for c in (7, 8, 11):
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True,
                                                           vertical='top')

    # DETALHADO
    montar(wb.create_sheet('DETALHADO'),
           ['CORTE', 'TRECHO', 'KM DO CORTE', 'CÂMERA', 'VÍDEO',
            'KM INICIAL DO VÍDEO', 'KM FINAL DO VÍDEO', 'KM COBERTOS DO CORTE',
            '% DO CORTE', 'DESVIO MEDIANO (m)', 'DIA', 'HORA INÍCIO (UTC)',
            'HORA FIM (UTC)'],
           [[d['corte'], d['rodovia'], d['trecho'], d['camera'], d['video'],
             d['km_ini_video'], d['km_fim_video'], d['km_cobertos'], d['pct'],
             d['desvio_mediano'], d['dia'], d['hora_ini'], d['hora_fim']]
            for d in relatorio['detalhes']],
           [12, 10, 20, 11, 14, 20, 20, 20, 12, 18, 8, 17, 15])

    # NÃO COBERTOS
    montar(wb.create_sheet('NÃO COBERTOS'),
           ['CORTE', 'TRECHO', 'KM INICIAL', 'KM FINAL', 'EXTENSÃO (km)',
            'OBSERVAÇÃO'],
           [[p['corte'], p['rodovia'], p['km_ini'], p['km_fim'],
             p['extensao_km'], p['motivo']] for p in relatorio['pendencias']],
           [12, 10, 14, 14, 14, 42])

    # GPX que não entraram em nenhum corte
    if relatorio['sobrando'] or relatorio['ignorados']:
        montar(wb.create_sheet('GPX NÃO APROVEITADOS'),
               ['ARQUIVO', 'VÍDEO', 'CÂMERA', 'TRECHO', 'KM INICIAL',
                'KM FINAL', 'MOTIVO'],
               [[s['arquivo'], s['video'], s['camera'], s['rodovia'],
                 s['km_ini'], s['km_fim'], s['motivo']]
                for s in relatorio['sobrando']] +
               [[g['arquivo'], None, None, None, None, None, g['motivo']]
                for g in relatorio['ignorados']],
               [34, 14, 11, 10, 14, 14, 46])

    # CRITÉRIOS
    ws = wb.create_sheet('CRITÉRIOS')
    r = relatorio['resumo']
    rods = '; '.join(f"{k} ({v['marcos']} marcos, {v['extensao_km']:.1f} km)"
                     for k, v in relatorio['rodovias'].items())
    for chave, valor in [
        ('CRITÉRIOS USADOS NA ANÁLISE', None),
        (None, None),
        ('Referências:',
         f'Marcos quilométricos: {ref_kmz or "KMZ enviado"} — {rods}. '
         f'Cortes esperados: {ref_plan or "planilha enviada"} '
         f'(colunas B, C, D e E).'),
        ('Eixo de quilometragem:',
         'Os marcos do KMZ são encadeados numa rota contínua por rodovia. '
         'Onde a numeração reinicia (divisa), as séries são unidas pelas '
         'pontas. A extensão de cada corte é a distância real ao longo dos '
         'marcos, não a diferença entre os números de km.'),
        ('Km ambíguo:',
         'Quando o mesmo km existe em mais de um ponto da rodovia, vale o '
         'menor trecho contíguo que liga o km inicial ao final.'),
        ('Cobertura:',
         f'Um ponto do GPX conta como "na rodovia" até {DIST_MAX_ROTA} m do '
         f'eixo. O trecho entre dois pontos consecutivos conta como '
         f'percorrido até o limite físico de {V_MAX * 3.6:.0f} km/h.'),
        ('Status FECHADO:',
         f'As duas câmeras cobrem, cada uma, pelo menos '
         f'{TOL_FECHADO * 100:.0f}% da extensão do corte.'),
        ('Status PARCIAL:',
         'Há cobertura, mas falta trecho ou falta uma das câmeras.'),
        ('Status NÃO FILMADO:',
         f'Menos de {TOL_PARCIAL * 100:.0f}% do corte tem vídeo.'),
        ('Buracos:',
         f'A aba NÃO COBERTOS lista as falhas a partir de {BURACO_MIN} m; '
         f'abaixo disso é ruído de GPS.'),
        ('Observação:',
         f'Traz o que foi coberto e o que falta, em faixas de km, no sentido '
         f'do corte (do km inicial para o final). Acima de {FAIXAS_NA_OBS} '
         f'faixas o texto resume o excedente — a lista completa fica na aba '
         f'NÃO COBERTOS.'),
        ('Câmera:',
         'INTERNO/EXTERNO vêm do nome da pasta ou do ZIP que contém o GPX.'),
        ('Ordem dos vídeos:',
         'Quando há mais de um vídeo na mesma célula, estão em ordem de '
         'gravação (horário GPS).'),
        (None, None),
        ('Totais:',
         f"{r['cortes']} cortes — {r['fechados']} fechados, "
         f"{r['parciais']} parciais, {r['nao_filmados']} não filmados, "
         f"{r['sem_referencia']} sem referência. "
         f"{r['gpx_lidos']} GPX aproveitados, {r['gpx_ignorados']} ignorados, "
         f"{r['gpx_sobrando']} fora dos cortes."),
    ]:
        ws.append([chave, valor])
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 110
    for linha in ws.iter_rows(min_row=1, max_col=1):
        linha[0].font = Font(bold=True)
    for linha in ws.iter_rows(min_row=1, min_col=2, max_col=2):
        linha[0].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A1'].font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
